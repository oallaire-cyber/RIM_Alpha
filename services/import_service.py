"""
Import Service.

Provides functionality to import RIM data from various formats,
primarily Excel spreadsheets.
"""

from typing import List, Dict, Any, Optional, Callable
from datetime import datetime
from dataclasses import dataclass, field
from config.settings import (
    RISK_LEVELS, RISK_STATUSES, RISK_ORIGINS, RISK_CATEGORIES,
    TPO_CLUSTERS, MITIGATION_TYPES, MITIGATION_STATUSES,
    INFLUENCE_STRENGTHS, IMPACT_LEVELS, MITIGATION_EFFECTIVENESS
)


@dataclass
class ImportResult:
    """Results from an import operation."""
    risks_created: int = 0
    risks_skipped: int = 0
    influences_created: int = 0
    influences_skipped: int = 0
    tpos_created: int = 0
    tpos_skipped: int = 0
    tpo_impacts_created: int = 0
    tpo_impacts_skipped: int = 0
    mitigations_created: int = 0
    mitigations_skipped: int = 0
    mitigates_created: int = 0
    mitigates_skipped: int = 0
    context_nodes_created: int = 0
    context_nodes_skipped: int = 0
    context_edges_created: int = 0
    context_edges_skipped: int = 0
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    logs: List[str] = field(default_factory=list)
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "risks_created": self.risks_created,
            "risks_skipped": self.risks_skipped,
            "influences_created": self.influences_created,
            "influences_skipped": self.influences_skipped,
            "tpos_created": self.tpos_created,
            "tpos_skipped": self.tpos_skipped,
            "tpo_impacts_created": self.tpo_impacts_created,
            "tpo_impacts_skipped": self.tpo_impacts_skipped,
            "mitigations_created": self.mitigations_created,
            "mitigations_skipped": self.mitigations_skipped,
            "mitigates_created": self.mitigates_created,
            "mitigates_skipped": self.mitigates_skipped,
            "context_nodes_created": self.context_nodes_created,
            "context_nodes_skipped": self.context_nodes_skipped,
            "context_edges_created": self.context_edges_created,
            "context_edges_skipped": self.context_edges_skipped,
            "errors": self.errors,
            "warnings": self.warnings,
            "logs": self.logs
        }
    
    def log(self, message: str, level: str = "INFO"):
        """Add a log entry with timestamp."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.logs.append(f"[{timestamp}] {level}: {message}")


class ExcelImporter:
    """
    Imports RIM data from Excel files.
    
    Handles validation, mapping, and creation of entities
    from Excel spreadsheet data.
    """
    
    def __init__(
        self,
        create_risk_fn: Callable,
        create_influence_fn: Callable,
        create_mitigation_fn: Callable,
        create_mitigates_fn: Callable,
        get_all_risks_fn: Callable,
        get_all_mitigations_fn: Callable,
        create_generic_entity_fn: Optional[Callable] = None,
        get_generic_entities_fn: Optional[Callable] = None,
        create_generic_relationship_fn: Optional[Callable] = None,
        registry=None,
    ):
        """
        Initialize the importer with database operation functions.
        
        Args:
            create_risk_fn: Function to create a risk
            create_influence_fn: Function to create an influence
            create_mitigation_fn: Function to create a mitigation
            create_mitigates_fn: Function to create a mitigates relationship
            get_all_risks_fn: Function to get all risks
            get_all_mitigations_fn: Function to get all mitigations
            create_generic_entity_fn: Function to create a ContextNode (type_id, data) -> dict
            get_generic_entities_fn: Function to get ContextNodes (type_id) -> list
            create_generic_relationship_fn: Function to create a ContextEdge
            registry: SchemaRegistry instance for ContextNode/ContextEdge type lookup
        """
        self.create_risk = create_risk_fn
        self.create_influence = create_influence_fn
        self.create_mitigation = create_mitigation_fn
        self.create_mitigates = create_mitigates_fn
        self.get_all_risks = get_all_risks_fn
        self.get_all_mitigations = get_all_mitigations_fn
        self.create_generic_entity = create_generic_entity_fn
        self.get_generic_entities = get_generic_entities_fn
        self.create_generic_relationship = create_generic_relationship_fn
        self.registry = registry
    
    def import_from_excel(self, filepath: str) -> ImportResult:
        """
        Import all data from an Excel file.
        
        Args:
            filepath: Path to the Excel file
        
        Returns:
            ImportResult with counts and logs
        """
        import pandas as pd
        
        result = ImportResult()
        result.log(f"Starting import from {filepath}")
        
        # Mappings for name-to-ID resolution
        risk_name_to_id: Dict[str, str] = {}
        tpo_ref_to_id: Dict[str, str] = {}
        mitigation_name_to_id: Dict[str, str] = {}
        
        try:
            # Import core entities
            self._import_risks(filepath, result)
            
            # Build risk name mapping
            result.log("Building risk name mapping from database...")
            all_risks = self.get_all_risks()
            for risk in all_risks:
                risk_name_to_id[risk['name']] = risk['id']
            result.log(f"Mapped {len(risk_name_to_id)} risks by name")
            
            # Import TPOs
            self._import_tpos(filepath, result)
            
            # Build TPO reference mapping
            result.log("Building TPO reference mapping from database...")
            all_tpos = self.get_all_tpos()
            for tpo in all_tpos:
                tpo_ref_to_id[tpo['reference']] = tpo['id']
            result.log(f"Mapped {len(tpo_ref_to_id)} TPOs by reference")
            
            # Import Influences
            self._import_influences(filepath, result, risk_name_to_id)
            
            # Import TPO Impacts
            self._import_tpo_impacts(filepath, result, risk_name_to_id, tpo_ref_to_id)
            
            # Import Mitigations
            self._import_mitigations(filepath, result)
            
            # Build mitigation name mapping
            result.log("Building mitigation name mapping from database...")
            all_mitigations = self.get_all_mitigations()
            for mit in all_mitigations:
                mitigation_name_to_id[mit['name']] = mit['id']
            result.log(f"Mapped {len(mitigation_name_to_id)} mitigations by name")
            
            # Import Mitigates relationships
            self._import_mitigates(filepath, result, mitigation_name_to_id, risk_name_to_id)
            
            # Import Context data (schema-driven, no new types created)
            if self.registry and self.create_generic_entity:
                self._import_context_nodes(filepath, result)
                # Build full node name map for edge resolution (risks + TPOs + context nodes)
                node_name_to_id: Dict[str, str] = {**risk_name_to_id}
                node_name_to_id.update(tpo_ref_to_id)
                if self.get_generic_entities:
                    for type_id in self.registry.entity_types:
                        if type_id in ("risk", "mitigation"):
                            continue
                        for entity in (self.get_generic_entities(type_id) or []):
                            if entity.get("name") and entity.get("id"):
                                node_name_to_id[entity["name"]] = entity["id"]
                if self.create_generic_relationship:
                    self._import_context_edges(filepath, result, node_name_to_id)
            
        except Exception as e:
            result.errors.append(f"Global import error: {str(e)}")
            result.log(f"Global error during import: {str(e)}", "ERROR")
        
        # Final summary
        result.log("=" * 50)
        result.log("Import completed:")
        result.log(f"  Risks: {result.risks_created} created, {result.risks_skipped} skipped")
        result.log(f"  TPOs: {result.tpos_created} created, {result.tpos_skipped} skipped")
        result.log(f"  Influences: {result.influences_created} created, {result.influences_skipped} skipped")
        result.log(f"  TPO Impacts: {result.tpo_impacts_created} created, {result.tpo_impacts_skipped} skipped")
        result.log(f"  Mitigations: {result.mitigations_created} created, {result.mitigations_skipped} skipped")
        result.log(f"  Mitigates: {result.mitigates_created} created, {result.mitigates_skipped} skipped")
        result.log(f"  Context Nodes: {result.context_nodes_created} created, {result.context_nodes_skipped} skipped")
        result.log(f"  Context Edges: {result.context_edges_created} created, {result.context_edges_skipped} skipped")
        result.log(f"  Errors: {len(result.errors)}, Warnings: {len(result.warnings)}")
        
        return result
    
    def _import_risks(self, filepath: str, result: ImportResult):
        """Import risks from Excel."""
        import pandas as pd
        
        result.log("Processing Risks sheet...")
        try:
            df = pd.read_excel(filepath, sheet_name='Risks')
            result.log(f"Found {len(df)} risks in Excel file")
            
            # Detect ext_* columns
            ext_columns = [c for c in df.columns if c.startswith("ext_")]
            if ext_columns:
                result.log(f"Found extension columns: {', '.join(ext_columns)}")
            
            for idx, row in df.iterrows():
                row_num = idx + 2
                try:
                    risk_name = row.get('name', 'Unknown')
                    
                    # Validate required fields
                    if pd.isna(row.get('name')):
                        result.warnings.append(f"Row {row_num}: Missing risk name, skipped")
                        result.risks_skipped += 1
                        continue
                    
                    if pd.isna(row.get('level')) or row.get('level') not in RISK_LEVELS:
                        result.warnings.append(f"Row {row_num} ({risk_name}): Invalid level '{row.get('level', 'None')}'. Valid: {RISK_LEVELS}")
                        result.risks_skipped += 1
                        continue
                    
                    # Parse categories
                    categories = self._parse_categories(row.get('categories', ['Programme']))
                    if not categories:
                        categories = ['Programme']
                        result.warnings.append(f"Row {row_num} ({risk_name}): Invalid categories, defaulting")
                    
                    # Parse other fields with defaults
                    description = self._safe_string(row.get('description', ''))
                    status = row.get('status', 'Active')
                    if pd.isna(status) or status not in RISK_STATUSES:
                        status = 'Active'
                    
                    origin = row.get('origin', 'New')
                    if pd.isna(origin) or origin not in RISK_ORIGINS:
                        origin = 'New'
                    
                    owner = self._safe_string(row.get('owner', ''))
                    probability = self._safe_float(row.get('probability'))
                    impact = self._safe_float(row.get('impact'))
                    
                    activation_condition = None
                    if not pd.isna(row.get('activation_condition')):
                        activation_condition = str(row.get('activation_condition'))
                    
                    activation_date = self._parse_date(row.get('activation_decision_date'))
                    
                    # Parse subtype
                    subtype = None
                    if 'subtype' in df.columns and not pd.isna(row.get('subtype')):
                        subtype = str(row.get('subtype'))
                    
                    # Parse extension fields
                    ext_fields = {}
                    for col in ext_columns:
                        val = row.get(col)
                        if not pd.isna(val):
                            # Cast booleans properly
                            if isinstance(val, (bool,)):
                                ext_fields[col] = val
                            elif isinstance(val, (int,)):
                                ext_fields[col] = int(val)
                            elif isinstance(val, (float,)):
                                ext_fields[col] = float(val)
                            else:
                                ext_fields[col] = str(val)
                    
                    # Create risk
                    if self.create_risk(
                        name=risk_name,
                        level=row['level'],
                        categories=categories,
                        description=description,
                        status=status,
                        activation_condition=activation_condition,
                        activation_decision_date=activation_date,
                        owner=owner,
                        probability=probability,
                        impact=impact,
                        origin=origin,
                        subtype=subtype,
                        ext_fields=ext_fields if ext_fields else None,
                    ):
                        result.risks_created += 1
                        result.log(f"Created risk: {risk_name}")
                    else:
                        result.risks_skipped += 1
                        result.warnings.append(f"Row {row_num} ({risk_name}): Failed to create")
                
                except Exception as e:
                    result.risks_skipped += 1
                    result.errors.append(f"Row {row_num} - Risk error: {str(e)}")
        
        except ValueError as e:
            if "Worksheet" in str(e):
                result.log("No 'Risks' sheet found", "WARNING")
            else:
                raise
        except Exception as e:
            result.errors.append(f"Risks sheet error: {str(e)}")
    
    def _import_tpos(self, filepath: str, result: ImportResult):
        """Import TPOs from Excel."""
        import pandas as pd
        
        result.log("Processing TPOs sheet...")
        try:
            df = pd.read_excel(filepath, sheet_name='TPOs')
            result.log(f"Found {len(df)} TPOs in Excel file")
            
            for idx, row in df.iterrows():
                row_num = idx + 2
                try:
                    tpo_ref = row.get('reference', 'Unknown')
                    
                    if pd.isna(row.get('reference')) or pd.isna(row.get('name')):
                        result.warnings.append(f"TPO Row {row_num}: Missing reference or name, skipped")
                        result.tpos_skipped += 1
                        continue
                    
                    cluster = row.get('cluster', 'Business Efficiency')
                    if pd.isna(cluster) or cluster not in TPO_CLUSTERS:
                        cluster = 'Business Efficiency'
                    
                    description = self._safe_string(row.get('description', ''))
                    
                    if self.create_tpo(
                        reference=row['reference'],
                        name=row['name'],
                        cluster=cluster,
                        description=description
                    ):
                        result.tpos_created += 1
                        result.log(f"Created TPO: {tpo_ref}")
                    else:
                        result.tpos_skipped += 1
                
                except Exception as e:
                    result.tpos_skipped += 1
                    result.errors.append(f"TPO Row {row_num} - Error: {str(e)}")
        
        except ValueError:
            result.log("No 'TPOs' sheet found", "WARNING")
        except Exception as e:
            result.log(f"TPOs sheet error: {str(e)}", "WARNING")
    
    def _import_influences(
        self,
        filepath: str,
        result: ImportResult,
        risk_name_to_id: Dict[str, str]
    ):
        """Import influences from Excel."""
        import pandas as pd
        
        result.log("Processing Influences sheet...")
        try:
            df = pd.read_excel(filepath, sheet_name='Influences')
            result.log(f"Found {len(df)} influences in Excel file")
            
            for idx, row in df.iterrows():
                row_num = idx + 2
                try:
                    source_name = row.get('source_name')
                    target_name = row.get('target_name')
                    
                    if pd.isna(source_name) or pd.isna(target_name):
                        result.warnings.append(f"Influence Row {row_num}: Missing names, skipped")
                        result.influences_skipped += 1
                        continue
                    
                    source_id = risk_name_to_id.get(source_name)
                    target_id = risk_name_to_id.get(target_name)
                    
                    if not source_id or not target_id:
                        result.warnings.append(f"Influence Row {row_num}: Risk not found, skipped")
                        result.influences_skipped += 1
                        continue
                    
                    strength = row.get('strength', 'Moderate')
                    if pd.isna(strength) or strength not in INFLUENCE_STRENGTHS:
                        strength = 'Moderate'
                    
                    confidence = self._safe_float(row.get('confidence'), 0.8)
                    description = self._safe_string(row.get('description', ''))
                    
                    if self.create_influence(
                        source_id=source_id,
                        target_id=target_id,
                        influence_type=row.get('influence_type', ''),
                        strength=strength,
                        description=description,
                        confidence=confidence
                    ):
                        result.influences_created += 1
                        result.log(f"Created influence: {source_name} → {target_name}")
                    else:
                        result.influences_skipped += 1
                
                except Exception as e:
                    result.influences_skipped += 1
                    result.errors.append(f"Influence Row {row_num} - Error: {str(e)}")
        
        except ValueError:
            result.log("No 'Influences' sheet found", "WARNING")
        except Exception as e:
            result.log(f"Influences sheet error: {str(e)}", "WARNING")
    
    def _import_tpo_impacts(
        self,
        filepath: str,
        result: ImportResult,
        risk_name_to_id: Dict[str, str],
        tpo_ref_to_id: Dict[str, str]
    ):
        """Import TPO impacts from Excel."""
        import pandas as pd
        
        result.log("Processing TPO_Impacts sheet...")
        try:
            df = pd.read_excel(filepath, sheet_name='TPO_Impacts')
            result.log(f"Found {len(df)} TPO impacts in Excel file")
            
            for idx, row in df.iterrows():
                row_num = idx + 2
                try:
                    risk_name = row.get('risk_name')
                    tpo_reference = row.get('tpo_reference')
                    
                    if pd.isna(risk_name) or pd.isna(tpo_reference):
                        result.warnings.append(f"TPO Impact Row {row_num}: Missing names, skipped")
                        result.tpo_impacts_skipped += 1
                        continue
                    
                    risk_id = risk_name_to_id.get(risk_name)
                    tpo_id = tpo_ref_to_id.get(tpo_reference)
                    
                    if not risk_id or not tpo_id:
                        result.warnings.append(f"TPO Impact Row {row_num}: Entity not found, skipped")
                        result.tpo_impacts_skipped += 1
                        continue
                    
                    impact_level = row.get('impact_level', 'Medium')
                    if pd.isna(impact_level) or impact_level not in IMPACT_LEVELS:
                        impact_level = 'Medium'
                    
                    description = self._safe_string(row.get('description', ''))
                    
                    if self.create_tpo_impact(
                        risk_id=risk_id,
                        tpo_id=tpo_id,
                        impact_level=impact_level,
                        description=description
                    ):
                        result.tpo_impacts_created += 1
                        result.log(f"Created TPO impact: {risk_name} → {tpo_reference}")
                    else:
                        result.tpo_impacts_skipped += 1
                
                except Exception as e:
                    result.tpo_impacts_skipped += 1
                    result.errors.append(f"TPO Impact Row {row_num} - Error: {str(e)}")
        
        except ValueError:
            result.log("No 'TPO_Impacts' sheet found", "WARNING")
        except Exception as e:
            result.log(f"TPO_Impacts sheet error: {str(e)}", "WARNING")
    
    def _import_mitigations(self, filepath: str, result: ImportResult):
        """Import mitigations from Excel."""
        import pandas as pd
        
        result.log("Processing Mitigations sheet...")
        try:
            df = pd.read_excel(filepath, sheet_name='Mitigations')
            result.log(f"Found {len(df)} mitigations in Excel file")
            
            for idx, row in df.iterrows():
                row_num = idx + 2
                try:
                    mit_name = row.get('name', 'Unknown')
                    
                    if pd.isna(row.get('name')):
                        result.warnings.append(f"Mitigation Row {row_num}: Missing name, skipped")
                        result.mitigations_skipped += 1
                        continue
                    
                    mit_type = row.get('type', 'Dedicated')
                    if pd.isna(mit_type) or mit_type not in MITIGATION_TYPES:
                        mit_type = 'Dedicated'
                    
                    mit_status = row.get('status', 'Proposed')
                    if pd.isna(mit_status) or mit_status not in MITIGATION_STATUSES:
                        mit_status = 'Proposed'
                    
                    description = self._safe_string(row.get('description', ''))
                    owner = self._safe_string(row.get('owner', ''))
                    source_entity = self._safe_string(row.get('source_entity', ''))
                    
                    if self.create_mitigation(
                        name=mit_name,
                        mitigation_type=mit_type,
                        status=mit_status,
                        description=description,
                        owner=owner,
                        source_entity=source_entity
                    ):
                        result.mitigations_created += 1
                        result.log(f"Created mitigation: {mit_name}")
                    else:
                        result.mitigations_skipped += 1
                
                except Exception as e:
                    result.mitigations_skipped += 1
                    result.errors.append(f"Mitigation Row {row_num} - Error: {str(e)}")
        
        except ValueError:
            result.log("No 'Mitigations' sheet found", "WARNING")
        except Exception as e:
            result.log(f"Mitigations sheet error: {str(e)}", "WARNING")
    
    def _import_mitigates(
        self,
        filepath: str,
        result: ImportResult,
        mitigation_name_to_id: Dict[str, str],
        risk_name_to_id: Dict[str, str]
    ):
        """Import mitigates relationships from Excel."""
        import pandas as pd
        
        result.log("Processing Mitigates sheet...")
        try:
            df = pd.read_excel(filepath, sheet_name='Mitigates')
            result.log(f"Found {len(df)} mitigates relationships in Excel file")
            
            for idx, row in df.iterrows():
                row_num = idx + 2
                try:
                    mitigation_name = row.get('mitigation_name')
                    risk_name = row.get('risk_name')
                    
                    if pd.isna(mitigation_name) or pd.isna(risk_name):
                        result.warnings.append(f"Mitigates Row {row_num}: Missing names, skipped")
                        result.mitigates_skipped += 1
                        continue
                    
                    mitigation_id = mitigation_name_to_id.get(mitigation_name)
                    risk_id = risk_name_to_id.get(risk_name)
                    
                    if not mitigation_id or not risk_id:
                        result.warnings.append(f"Mitigates Row {row_num}: Entity not found, skipped")
                        result.mitigates_skipped += 1
                        continue
                    
                    effectiveness = row.get('effectiveness', 'Medium')
                    if pd.isna(effectiveness) or effectiveness not in MITIGATION_EFFECTIVENESS:
                        effectiveness = 'Medium'
                    
                    description = self._safe_string(row.get('description', ''))
                    
                    if self.create_mitigates(
                        mitigation_id=mitigation_id,
                        risk_id=risk_id,
                        effectiveness=effectiveness,
                        description=description
                    ):
                        result.mitigates_created += 1
                        result.log(f"Created mitigates: {mitigation_name} → {risk_name}")
                    else:
                        result.mitigates_skipped += 1
                
                except Exception as e:
                    result.mitigates_skipped += 1
                    result.errors.append(f"Mitigates Row {row_num} - Error: {str(e)}")
        
        except ValueError:
            result.log("No 'Mitigates' sheet found", "WARNING")
        except Exception as e:
            result.log(f"Mitigates sheet error: {str(e)}", "WARNING")
    
    def _parse_categories(self, value) -> List[str]:
        """Parse categories from various formats."""
        import pandas as pd
        import ast
        
        if pd.isna(value):
            return []
        
        if isinstance(value, list):
            return value
        
        if isinstance(value, str):
            try:
                parsed = ast.literal_eval(value)
                if isinstance(parsed, list):
                    return parsed
            except:
                pass
            
            # Try comma-separated
            return [c.strip().strip("[]'\"") for c in value.split(',')]
        
        return []
    
    def _safe_string(self, value, default: str = '') -> str:
        """Safely convert value to string."""
        import pandas as pd
        if pd.isna(value):
            return default
        return str(value)
    
    def _safe_float(self, value, default: Optional[float] = None) -> Optional[float]:
        """Safely convert value to float."""
        import pandas as pd
        if pd.isna(value):
            return default
        try:
            return float(value)
        except (ValueError, TypeError):
            return default
    
    def _parse_date(self, value) -> Optional[str]:
        """Parse date value to ISO string."""
        import pandas as pd
        if pd.isna(value):
            return None
        if hasattr(value, 'isoformat'):
            return value.isoformat()
        return str(value) if value else None

    def _import_context_nodes(self, filepath: str, result: ImportResult) -> None:
        """
        Import all ContextNode sheets (sheets prefixed with 'CN_') from an Excel file.

        For each 'CN_{type_id}' sheet found:
        - If the type_id is not in the schema registry, the sheet is skipped and a
          structured [SCHEMA] warning with a ready-to-paste YAML snippet is logged.
        - Extra columns not declared in the schema property list emit a one-time
          per-sheet [SCHEMA] warning pointing to the exact schema.yaml location.
        """
        import pandas as pd
        import openpyxl

        result.log("Processing Context Node sheets (CN_*) ...")

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            result.log(f"Could not open workbook for context nodes: {e}", "WARNING")
            return

        cn_sheets = [s for s in wb.sheetnames if s.startswith("CN_")]
        result.log(f"Found {len(cn_sheets)} context node sheet(s): {cn_sheets or 'none'}")

        for sheet_name in cn_sheets:
            type_id = sheet_name[3:]  # strip "CN_"

            # --- Registry check ---
            entity_type = self.registry.get_entity_type(type_id) if self.registry else None
            if entity_type is None or type_id in ("risk", "mitigation"):
                # Read column headers for the YAML scaffold
                ws = wb[sheet_name]
                col_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                col_headers = [c for c in col_headers if c]
                props_yaml = "\n".join(
                    f"            - {{ name: \"{c}\", type: \"string\" }}" for c in col_headers
                )
                warning = (
                    f"\n[SCHEMA] Sheet '{sheet_name}' skipped — type '{type_id}' not found in schema.yaml.\n"
                    f"To enable this import, add the following to schema.yaml under 'context_nodes':\n\n"
                    f"    {type_id}:\n"
                    f"      shape: \"box\"\n"
                    f"      color: \"#AAAAAA\"\n"
                    f"      zone: \"lower\"   # or \"upper\"\n"
                    f"      properties:\n"
                    f"{props_yaml}\n\n"
                    f"Then re-import the file.\n"
                )
                result.warnings.append(warning)
                result.log(f"[SCHEMA] Sheet '{sheet_name}' skipped (type unknown)", "WARNING")
                # Count all rows (minus header) as skipped
                row_count = sum(1 for _ in ws.iter_rows(min_row=2)) - 1
                result.context_nodes_skipped += max(row_count, 0)
                continue

            # --- Load sheet with pandas for easy row iteration ---
            try:
                df = pd.read_excel(filepath, sheet_name=sheet_name)
            except Exception as e:
                result.log(f"Sheet '{sheet_name}': could not read — {e}", "WARNING")
                continue

            # --- Column mismatch check ---
            schema_prop_names = {p.name for p in entity_type.properties}
            sheet_cols = set(df.columns.tolist())
            unknown_cols = sheet_cols - schema_prop_names - {"name"}
            if unknown_cols:
                loc = f"context_nodes.{type_id}.properties"
                result.warnings.append(
                    f"\n[SCHEMA] Sheet '{sheet_name}': unrecognized columns {sorted(unknown_cols)} ignored.\n"
                    f"These are not declared in schema.yaml under '{loc}'.\n"
                    f"To persist them, add entries like:\n"
                    + "\n".join(
                        f"    - {{ name: \"{c}\", type: \"string\" }}"
                        for c in sorted(unknown_cols)
                    )
                    + f"\nto the '{loc}' list.\n"
                )
                result.log(f"[SCHEMA] Sheet '{sheet_name}': {len(unknown_cols)} unknown column(s) ignored", "WARNING")

            # --- Import rows ---
            result.log(f"Importing '{sheet_name}' ({len(df)} rows) ...")
            for idx, row in df.iterrows():
                row_num = idx + 2
                try:
                    # Build data dict using only schema-declared properties (+ name)
                    data: Dict[str, Any] = {}
                    if "name" in df.columns and not pd.isna(row.get("name")):
                        data["name"] = str(row["name"])
                    elif "name" in df.columns:
                        result.warnings.append(f"[{sheet_name}] Row {row_num}: missing 'name', skipped")
                        result.context_nodes_skipped += 1
                        continue

                    for prop in entity_type.properties:
                        if prop.name in df.columns:
                            val = row.get(prop.name)
                            if not pd.isna(val):
                                data[prop.name] = val

                    created = self.create_generic_entity(type_id, data)
                    if created:
                        result.context_nodes_created += 1
                        result.log(f"Created {type_id}: {data.get('name', '?')}")
                    else:
                        result.context_nodes_skipped += 1

                except Exception as e:
                    result.context_nodes_skipped += 1
                    result.errors.append(f"[{sheet_name}] Row {row_num}: {str(e)}")

        wb.close()

    def _import_context_edges(
        self,
        filepath: str,
        result: ImportResult,
        node_name_to_id: Dict[str, str],
    ) -> None:
        """
        Import all ContextEdge sheets (sheets prefixed with 'CE_') from an Excel file.

        For each 'CE_{rel_type_id}' sheet found:
        - If the rel_type_id is not in the schema registry, the sheet is skipped and a
          structured [SCHEMA] warning with a ready-to-paste YAML snippet is logged.
        - Extra columns not declared in the schema property list emit a one-time
          per-sheet [SCHEMA] warning.
        - Rows with unresolvable source_name / target_name are skipped with a warning.
        """
        import pandas as pd
        import openpyxl

        result.log("Processing Context Edge sheets (CE_*) ...")

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            result.log(f"Could not open workbook for context edges: {e}", "WARNING")
            return

        ce_sheets = [s for s in wb.sheetnames if s.startswith("CE_")]
        result.log(f"Found {len(ce_sheets)} context edge sheet(s): {ce_sheets or 'none'}")

        kernel_rel_ids = {"influences", "mitigates"}

        for sheet_name in ce_sheets:
            rel_type_id = sheet_name[3:]  # strip "CE_"

            # --- Registry check ---
            rel_type = self.registry.get_relationship_type(rel_type_id) if self.registry else None
            if rel_type is None or rel_type_id in kernel_rel_ids:
                ws = wb[sheet_name]
                col_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                col_headers = [c for c in col_headers if c and c not in ("source_name", "target_name")]
                props_yaml = "\n".join(
                    f"            - {{ name: \"{c}\", type: \"string\" }}" for c in col_headers
                ) if col_headers else "            []"
                warning = (
                    f"\n[SCHEMA] Sheet '{sheet_name}' skipped — relationship type '{rel_type_id}' not found in schema.yaml.\n"
                    f"To enable this import, add the following to schema.yaml under 'relationship_types':\n\n"
                    f"    {rel_type_id}:\n"
                    f"      semantic: \"context\"          # or \"influence\" / \"cluster\"\n"
                    f"      valid_from: [\"<source_node_type>\"]\n"
                    f"      valid_to:   [\"<target_node_type>\"]\n"
                    f"      properties:\n"
                    f"{props_yaml}\n\n"
                    f"Then re-import the file.\n"
                )
                result.warnings.append(warning)
                result.log(f"[SCHEMA] Sheet '{sheet_name}' skipped (rel type unknown)", "WARNING")
                row_count = sum(1 for _ in ws.iter_rows(min_row=2)) - 1
                result.context_edges_skipped += max(row_count, 0)
                continue

            # --- Load sheet ---
            try:
                df = pd.read_excel(filepath, sheet_name=sheet_name)
            except Exception as e:
                result.log(f"Sheet '{sheet_name}': could not read — {e}", "WARNING")
                continue

            # --- Column mismatch check ---
            schema_prop_names = {p.name for p in rel_type.properties}
            reserved = {"source_name", "target_name"}
            sheet_cols = set(df.columns.tolist())
            unknown_cols = sheet_cols - schema_prop_names - reserved
            if unknown_cols:
                loc = f"relationship_types.{rel_type_id}.properties"
                result.warnings.append(
                    f"\n[SCHEMA] Sheet '{sheet_name}': unrecognized columns {sorted(unknown_cols)} ignored.\n"
                    f"These are not declared in schema.yaml under '{loc}'.\n"
                    f"To persist them, add entries like:\n"
                    + "\n".join(
                        f"    - {{ name: \"{c}\", type: \"string\" }}"
                        for c in sorted(unknown_cols)
                    )
                    + f"\nto the '{loc}' list.\n"
                )
                result.log(f"[SCHEMA] Sheet '{sheet_name}': {len(unknown_cols)} unknown column(s) ignored", "WARNING")

            # --- Import rows ---
            result.log(f"Importing '{sheet_name}' ({len(df)} rows) ...")
            for idx, row in df.iterrows():
                row_num = idx + 2
                try:
                    source_name = row.get("source_name")
                    target_name = row.get("target_name")

                    if pd.isna(source_name) or pd.isna(target_name):
                        result.warnings.append(f"[{sheet_name}] Row {row_num}: missing source_name or target_name, skipped")
                        result.context_edges_skipped += 1
                        continue

                    source_id = node_name_to_id.get(str(source_name))
                    target_id = node_name_to_id.get(str(target_name))

                    if not source_id:
                        result.warnings.append(f"[{sheet_name}] Row {row_num}: source '{source_name}' not found in graph, skipped")
                        result.context_edges_skipped += 1
                        continue
                    if not target_id:
                        result.warnings.append(f"[{sheet_name}] Row {row_num}: target '{target_name}' not found in graph, skipped")
                        result.context_edges_skipped += 1
                        continue

                    # Build property data dict (schema-declared props only)
                    data: Dict[str, Any] = {}
                    for prop in rel_type.properties:
                        if prop.name in df.columns:
                            val = row.get(prop.name)
                            if not pd.isna(val):
                                data[prop.name] = val

                    created = self.create_generic_relationship(
                        rel_type_id, source_id, target_id, data
                    )
                    if created:
                        result.context_edges_created += 1
                        result.log(f"Created {rel_type_id}: {source_name} → {target_name}")
                    else:
                        result.context_edges_skipped += 1

                except Exception as e:
                    result.context_edges_skipped += 1
                    result.errors.append(f"[{sheet_name}] Row {row_num}: {str(e)}")

        wb.close()

